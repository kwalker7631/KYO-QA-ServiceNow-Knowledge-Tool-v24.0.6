import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parents[1]))

import openpyxl
from openpyxl.styles import Alignment
from excel_generator import generate_excel


def test_excel_formatting_preserved(tmp_path):
    template = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Page 1"
    ws.append(["needs_review", "Description"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    for cell in ws[1]:
        cell.alignment = Alignment(wrap_text=True)
    wb.save(template)

    data = [{"needs_review": True, "Description": "Some long text"}]
    output = tmp_path / "out.xlsx"
    generate_excel(data, output, template)

    res_wb = openpyxl.load_workbook(output)
    res_ws = res_wb["Page 1"]

    assert res_ws.column_dimensions["A"].width > 8.4
    assert res_ws.column_dimensions["B"].width > 8.4
    assert all(cell.alignment.wrap_text for cell in res_ws[1])

    fill = res_ws[2][0].fill
    assert fill.start_color.rgb.endswith("FFF3BF")

