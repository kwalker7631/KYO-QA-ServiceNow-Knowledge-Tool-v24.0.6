import sys
from pathlib import Path
sys.path.append(str(Path(__file__).resolve().parents[1]))
import openpyxl
from excel_generator import generate_excel


def test_generate_excel_writes_rows(tmp_path):
    headers = ["Active", "Article type", "Author"]
    template = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Page 1"
    ws.append(headers)
    wb.save(template)

    data = [{"Active": "TRUE", "Article type": "HTML", "Author": "Tester"}]
    output = tmp_path / "out.xlsx"
    generate_excel(data, output, template)

    result_wb = openpyxl.load_workbook(output)
    result_ws = result_wb["Page 1"]

    assert result_ws.max_row == 2
    assert [cell.value for cell in result_ws[2]] == ["TRUE", "HTML", "Tester"]


def test_generate_excel_formats_rows(tmp_path):
    headers = ["Active", "Article type", "Author"]
    template = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Page 1"
    ws.append(headers)
    wb.save(template)

    data = [
        {"Active": "TRUE", "Article type": "HTML", "Author": "Tester", "needs_review": True},
        {"Active": "FALSE", "Article type": "PDF", "Author": "Agent", "status": "failed"},
    ]
    output = tmp_path / "out.xlsx"
    generate_excel(data, output, template)

    result_wb = openpyxl.load_workbook(output)
    result_ws = result_wb["Page 1"]

    assert result_ws.column_dimensions["A"].width >= len("Active")
    assert all(cell.alignment.wrap_text for cell in result_ws[2])
    first_row_colors = {cell.fill.start_color.rgb for cell in result_ws[2]}
    second_row_colors = {cell.fill.start_color.rgb for cell in result_ws[3]}
    assert first_row_colors == {"00FFEB9C"}
    assert second_row_colors == {"00FFC7CE"}
