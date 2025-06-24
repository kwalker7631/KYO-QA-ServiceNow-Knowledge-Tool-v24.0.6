# KYO QA ServiceNow Excel Generator v24.0.6
import pandas as pd, shutil, openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill
from logging_utils import setup_logger, log_info, log_error, log_warning
from custom_exceptions import ExcelGenerationError

logger = setup_logger("excel_generator")

# Default cell styling
REVIEW_FILL = PatternFill(start_color="FFF3BF", end_color="FFF3BF", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

def _apply_formatting(sheet: openpyxl.worksheet.worksheet.Worksheet):
    """Auto-fit columns, wrap text, and apply row highlights."""
    # Track max character length per column for auto-fit
    max_width = {}

    status_idx = None
    review_idx = None
    for idx, cell in enumerate(sheet[1], start=1):
        header = str(cell.value).lower()
        if header == "needs_review":
            review_idx = idx
        elif header == "status":
            status_idx = idx

    # Examine all cells to calculate optimal column widths
    for row in sheet.iter_rows():
        for cell in row:
            text = str(cell.value) if cell.value is not None else ""
            letter = cell.column_letter
            max_width[letter] = max(len(text), max_width.get(letter, 0))
            if cell.row > 1:
                cell.alignment = Alignment(wrap_text=True)

    for letter, length in max_width.items():
        sheet.column_dimensions[letter].width = min(max(length + 2, 15), 50)

    # Apply conditional coloring for each data row
    for row in sheet.iter_rows(min_row=2):
        row_fill = None
        if review_idx and str(row[review_idx - 1].value).lower() in {"true", "needs_review", "1"}:
            row_fill = REVIEW_FILL
        elif status_idx:
            status_value = str(row[status_idx - 1].value).lower()
            if status_value in {"error", "failed"}:
                row_fill = ERROR_FILL
            elif status_value in {"completed", "ok", "success"}:
                row_fill = SUCCESS_FILL

        if row_fill:
            for cell in row:
                cell.fill = row_fill

def _use_template_excel(df, output_path, template_path):
    """Write DataFrame to an Excel file based on an existing template."""
    try:
        shutil.copy2(template_path, output_path)
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook["Page 1"] if "Page 1" in workbook.sheetnames else workbook.active
        for row in dataframe_to_rows(df, index=False, header=False):
            sheet.append(row)
        _apply_formatting(sheet)
        workbook.save(output_path)
        log_info(logger, f"Successfully saved data to {output_path}")
        return str(output_path)
    except Exception as e:
        raise ExcelGenerationError(f"Failed to write data to template: {e}")


def _create_new_excel(df, output_path):
    """Write DataFrame to a brand new Excel file when no template is provided."""
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)
        _apply_formatting(sheet)
        workbook.save(output_path)
        log_info(logger, f"Successfully saved data to {output_path}")
        return str(output_path)
    except Exception as e:
        raise ExcelGenerationError(f"Failed to write data: {e}")

def generate_excel(all_results, output_path, template_path):
    try:
        if not all_results: raise ExcelGenerationError("No data to generate.")
        df = pd.DataFrame(all_results)
        
        if template_path:
            log_info(logger, f"Aligning data with template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook["Page 1"] if "Page 1" in workbook.sheetnames else workbook.active
            template_headers = [cell.value for cell in sheet[1] if cell.value]
            workbook.close()
            if not template_headers:
                raise ExcelGenerationError("No headers in template.")

            extra_cols = [col for col in df.columns if col not in template_headers]
            if extra_cols:
                log_warning(logger, f"Dropping non-template columns: {extra_cols}")
                df = df.drop(columns=extra_cols, errors='ignore')

            missing_cols = [col for col in template_headers if col not in df.columns]
            if missing_cols:
                log_warning(logger, f"Adding missing template columns: {missing_cols}")
                for col in missing_cols:
                    df[col] = ""

            df = df[template_headers]
            return _use_template_excel(df, output_path, template_path)
        else:
            log_info(logger, "No template provided. Creating new workbook.")
            df.to_excel(output_path, index=False)
            workbook = openpyxl.load_workbook(output_path)
            sheet = workbook.active
            _apply_formatting(sheet)
            workbook.save(output_path)
            log_info(logger, f"Successfully saved data to {output_path}")
            return str(output_path)
    except Exception as e:
        log_error(logger, f"Excel generation failed: {e}")
        raise ExcelGenerationError(f"Failed to generate Excel file: {e}")
