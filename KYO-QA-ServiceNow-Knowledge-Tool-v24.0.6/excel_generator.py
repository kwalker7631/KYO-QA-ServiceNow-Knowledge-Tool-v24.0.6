# KYO QA ServiceNow Excel Generator - FINAL VERSION with Conditional Formatting
from version import VERSION
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill
import re
from logging_utils import setup_logger, log_info, log_error
from custom_exceptions import ExcelGenerationError

logger = setup_logger("excel_generator")

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010\013\014\016-\037]')

# --- NEW: Define Fill Colors for Conditional Formatting ---
NEEDS_REVIEW_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
OCR_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")          # Light Yellow
FAILED_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")        # Dark Red
SUCCESS_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="none")       # No Fill

# This is the default structure, based on your latest template
DEFAULT_TEMPLATE_HEADERS = [
    'Active', 'Article type', 'Author', 'Category(category)', 'Configuration item',
    'Confidence', 'Description', 'Attachment link', 'Disable commenting',
    'Disable suggesting', 'Display attachments', 'Flagged', 'Governance',
    'Category(kb_category)', 'Knowledge base', 'Meta', 'Meta Description',
    'Ownership Group', 'Published', 'Scheduled publish date', 'Short description',
    'Article body', 'Topic', 'Problem Code', 'models', 'Ticket#', 'Valid to',
    'View as allowed', 'Wiki', 'Sys ID', 'file_name', 'Change Type', 'Revision'
]

def sanitize_for_excel(value):
    if isinstance(value, str): return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value

def apply_excel_styles(worksheet, df):
    """
    Applies word wrap, auto-fit columns, AND conditional row coloring.
    """
    log_info(logger, "Applying formatting and conditional coloring...")
    
    header_font = Font(bold=True)
    status_colors = {
        'Needs Review': NEEDS_REVIEW_FILL,
        'OCR Required': OCR_FILL,
        'Failed': FAILED_FILL,
        'Success': SUCCESS_FILL
    }
    
    # Iterate through the data rows to apply conditional formatting
    for index, data_row in df.iterrows():
        # Get the status for the current row
        status = data_row.get('processing_status', 'Success')
        fill_color = status_colors.get(status, SUCCESS_FILL)
        
        # Worksheet rows are 1-based, plus 1 for the header
        worksheet_row_index = index + 2
        
        for cell in worksheet[worksheet_row_index]:
            cell.fill = fill_color

    # Apply text formatting and auto-size columns
    for row_index, row in enumerate(worksheet.iter_rows()):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            if row_index == 0:
                cell.font = header_font

    for column_cells in worksheet.columns:
        try:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min((max_length + 2), 80)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        except Exception as e:
            log_error(logger, f"Could not set width for column {column_letter}: {e}")
            
    log_info(logger, "Worksheet styling complete.")

def generate_excel(all_results, output_path, template_path):
    try:
        if not all_results: raise ExcelGenerationError("No data to generate.")
        
        df = pd.DataFrame(all_results)
        df_sanitized = df.map(sanitize_for_excel)
        
        # Use default headers as the base
        headers = DEFAULT_TEMPLATE_HEADERS
        # If user provides a template, we can use its headers, but default is robust
        if template_path:
            log_info(logger, "User template provided. Aligning to default structure for stability.")

        final_df = pd.DataFrame(columns=headers)
        for col in headers:
            if col in df_sanitized.columns:
                final_df[col] = df_sanitized[col]
            else:
                final_df[col] = ""

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='ServiceNow Import')
            worksheet = writer.sheets['ServiceNow Import']
            # Pass the original dataframe with the status column to the styler
            apply_excel_styles(worksheet, df_sanitized)
        
        log_info(logger, f"Successfully created formatted Excel file: {output_path}")
        return str(output_path)
        
    except Exception as e:
        log_error(logger, f"Excel generation failed: {e}")
        raise ExcelGenerationError(f"Failed to generate Excel file: {e}")